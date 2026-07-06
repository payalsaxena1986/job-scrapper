#!/usr/bin/env python3
"""Build the daily job_matches_YYYY-MM-DD.xlsx workbook from a JSON list of roles.

Usage: python3 build_job_matches_excel.py <input.json> <output.xlsx>

Input JSON is a list of objects with keys matching COLUMNS (see below).
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Company", 18),
    ("Role name", 30),
    ("Job link", 32),
    ("Location", 20),
    ("Source", 16),
    ("Posted date or freshness evidence", 26),
    ("Probability of shortlist", 14),
    ("Reason for score", 42),
    ("Outreach plan", 34),
    ("Tailored resume headline", 34),
    ("Resume bullets to emphasize", 50),
    ("Keywords to include", 30),
    ("Gaps or risks", 34),
]

SCORE_FILL = {
    "High": PatternFill("solid", fgColor="C6EFCE"),
    "Medium": PatternFill("solid", fgColor="FFEB9C"),
    "Low": PatternFill("solid", fgColor="FFC7CE"),
}


def build(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Matches"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (name, _width) in enumerate(COLUMNS, start=1):
            value = row.get(name, "")
            if isinstance(value, list):
                value = "\n".join(f"- {v}" for v in value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = wrap_top
            if name == "Probability of shortlist" and value in SCORE_FILL:
                cell.fill = SCORE_FILL[value]
                cell.font = Font(bold=True)
        ws.row_dimensions[r_idx].height = 110

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    wb.save(out_path)
    print(f"Wrote {len(rows)} rows to {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    with open(sys.argv[1]) as f:
        rows = json.load(f)
    build(rows, sys.argv[2])
